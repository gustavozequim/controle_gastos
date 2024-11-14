from openpyxl import Workbook
from datetime import datetime as dt

from src.settings import OPCOES_ACEITAS_DE_GASTOS

class CriaPlanilhaControle:
    def cria_planilha(dicionario_de_gastos):
        opcao = str(input('Deseja criar a planilha de controle?[S/N]: ')).upper()
        while opcao not in OPCOES_ACEITAS_DE_GASTOS:
            print("Por favor digite S para inserir outro gasto, ou N para finalizar!")
            opcao = input('Deseja criar a planilha de controle?[S/N]: ').upper()
            if opcao == 'N':
                print('Obrigado por usar nossos serviços!')
                break
        if opcao == 'S':
            data = dt.now().strftime("%d-%m-%Y")
            wb = Workbook()
            planilha = wb.worksheets[0]
            planilha.title = "NOVEMBRO 2024"
            planilha['A1'] = "Tipo gasto"
            planilha['B1'] = "Valor"
            planilha['C1'] = "Data"

            for chave, valor in dicionario_de_gastos.items():
                planilha.cell(column=1, row=planilha.max_row + 1, value=chave)
                planilha.cell(column=2, row=planilha.max_row, value=valor)
                planilha.cell(column=3, row=planilha.max_row, value=data)
            wb.save('Controle_gastos.xlsx')
        print('\nPlanilha criada com sucesso')
