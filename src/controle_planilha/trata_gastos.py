import os
from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import datetime as dt

from src.settings import OPCOES_ACEITAS_DE_GASTOS, OPCOES_PARA_CONFIGURAR_PLANILHA

class PlanilhaControle:
    
    def cria_planilha(dicionario_de_gastos):
        opcao = str(input('Deseja criar a planilha de controle?[S/N]: ')).upper()
        while opcao not in OPCOES_ACEITAS_DE_GASTOS:
            print("Por favor digite S para inserir outro gasto, ou N para finalizar!")
            opcao = input('Deseja criar a planilha de controle?[S/N]: ').upper()
            if opcao == 'N':
                print('Obrigado por usar nossos serviços!')
                break
        nome_pagina = str(input("Digite o nome da página da planilha: "))
        nome_planilha = str(input("Digite o nome da planilha: "))
        if opcao == 'S':
            data = dt.now().strftime("%d-%m-%Y")
            wb = Workbook()
            planilha = wb.worksheets[0]
            planilha.title = nome_pagina
            planilha['A1'] = "Tipo gasto"
            planilha['B1'] = "Valor"
            planilha['C1'] = "Data"

            for chave, valor in dicionario_de_gastos.items():
                planilha.cell(column=1, row=planilha.max_row + 1, value=chave)
                planilha.cell(column=2, row=planilha.max_row, value=valor)
                planilha.cell(column=3, row=planilha.max_row, value=data)
            wb.save(f'{nome_planilha}.xlsx')
        print('\nPlanilha criada com sucesso')

    def deleta_dado(nome_planilha):
        wb = load_workbook(f'{nome_planilha}.xlsx')
        planilha = wb.worksheets[0]
        intervalo = planilha.iter_rows(max_row=planilha.max_row, max_col=planilha.max_column)
        intervalo = [[item.value for item in linha] for linha in intervalo]
        df = pd.DataFrame(intervalo)
        df.rename(columns=df.iloc[0], inplace=True)
        df.drop([0], inplace=True)
        print("Atualmente a planilha contém os seguintes dados")
        print(df)
        opcao_dado = int(input("Qual linha deseja excluir?: "))
        planilha.delete_rows(opcao_dado + 1)
        wb.save(f'{nome_planilha}.xlsx')
        print("Dado deletado com sucesso.")
    
    def adiciona_dados(nome_planilha, dicionario_de_gastos):
        wb = load_workbook(f'{nome_planilha}.xlsx')
        data = dt.now().strftime("%d-%m-%Y")
        planilha = wb.active
        planilha['A1'] = "Tipo gasto"
        planilha['B1'] = "Valor"
        planilha['C1'] = "Data"

        for chave, valor in dicionario_de_gastos.items():
            planilha.cell(column=1, row=planilha.max_row + 1, value=chave)
            planilha.cell(column=2, row=planilha.max_row, value=valor)
            planilha.cell(column=3, row=planilha.max_row, value=data)
        wb.save(f'{nome_planilha}.xlsx')
        print('\nPlanilha atualizada com sucesso')

    def cria_pagina(nome_planilha):
        wb = load_workbook(f'{nome_planilha}.xlsx')
        nome_pagina = str(input("Digite a página que deseja inserir: "))
        wb.create_sheet(nome_pagina, 0)
        wb.save(f'{nome_planilha}.xlsx')
        print(f'A página {nome_pagina} foi adicionada!')

    @classmethod
    def editar_planilha(cls, valores=None, nome_planilha=None):
        print("Você está no controle de planilhas")
        print("-="*30)
        continuar = True
        while continuar == True:
            print("-="*30)
            for config in OPCOES_PARA_CONFIGURAR_PLANILHA:
                print(f'{config}')
            print('-='*30)
            opcao = int(input("Digite o que deseja fazer: "))
            if opcao == 0:
                decisao = int(input("Deseja:\n[0] Criar Planilha\n[1] Atualizar Planilha\n"))
                if decisao == 0:
                    cls.cria_planilha(valores)
                if decisao == 1:
                    nome_planilha = str(input("Digite o nome da planilha para carregar: "))
                    cls.adiciona_dados(nome_planilha, valores)
            if opcao == 1:
                nome_planilha = str(input("Digite o nome da planilha para carregar: "))
                cls.deleta_dado(nome_planilha)
            if opcao == 2:
                nome_planilha = str(input("Digite o nome da planilha para carregar: "))
                cls.cria_pagina(nome_planilha)
            if opcao == 3:
                continuar = False
                print("Configuração de planilhas finalizado")
